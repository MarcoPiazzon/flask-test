from contatto import contatto_bp
from contatto import message
from flask import Blueprint,render_template, url_for, redirect, request
from flask_login import *
from home import *
from model import *
import openpyxl
from datetime import date

@contatto_bp.route('/<int:id>/', methods=['POST', 'GET'])
def home(id):
    print("sono nwellA HOM")
    current_contatto = None

    contatti = None
    # valore di default, indica che non sto cercando nessun utente
    try:
        if(id != 0):
            current_contatto = conn.execute(select(contatto).where(contatto.c.idcontatto == id and contatto.c.idutente == current_user.get_id())).fetchone()
        contatti = conn.execute(select(contatto).where(contatto.c.idutente == current_user.get_id()).order_by(contatto.c.nome)).fetchall()
    except Exception as error:
        conn.rollback()
        print("rip")
        print(error)
        print(error.__cause__)
    print(current_contatto)
    print(message)
    
    return render_template("/contatto/contatto.html", contatti=contatti, current_contatto=current_contatto, mess = message)
    

@contatto_bp.route('/getContatto', methods=['POST'])
def getContatto():
    
    idContatto = request.form['idSearch']
    print(request.form)
    print(idContatto)
    global message
    message = "PRESO IL CONTAT"
    return redirect(url_for('.home', id=idContatto))

def filterNumber(val):
    if(val is None):
        return None
    
    if(val is String):
        sub = sub.replace(" ", "")
        sub = list(val)
        if(len(sub) == 13 and list(val)[0] == '+'): # esempio '+393453659562' --> 3453659562
            return val[2:]
    
    return str(val)
    
@contatto_bp.route('/addContatto', methods=['POST'])
@login_required
def addContatto():  
    nome = request.form['nomeAdd']
    secondoNome = request.form['secondoNomeAdd']
    cognome = request.form['cognomeAdd']
    viaUfficio1 = request.form['viaUfficio1Add'] 
    viaUfficio2 = request.form['viaUfficio2Add']   
    viaUfficio3 = request.form['viaUfficio3Add']
    provincia = request.form['provinciaAdd']
    cap = request.form['capAdd']
    numUfficio = request.form['numUfficioAdd']
    numUfficio2 = request.form['numUfficio2Add']
    telefonoPrincipale = request.form['telefonoPrincipaleAdd']
    faxAbitazione = request.form['faxAbitazioneAdd']
    abitazione = request.form['abitazioneAdd']
    cellulare = request.form['cellulareAdd']
    note = request.form['noteAdd']
    email1 = request.form['email1Add']
    email2 = request.form['email2Add']
    paginaWeb = request.form['paginaWebAdd']
    id = 0
    try:
        ris = conn.execute(insert(contatto).values(
            nome = nome,
            idutente = current_user.get_id(),
            secondonome = secondoNome,
            cognome = cognome,
            viaufficio1 = viaUfficio1,
            viaufficio2 = viaUfficio2,
            viaufficio3 = viaUfficio3,
            provincia = provincia,
            cap = cap,
            numufficio = numUfficio,
            numufficio2 = numUfficio2,
            telefonoprincipale = telefonoPrincipale,
            faxabitazione = faxAbitazione,
            abitazione = abitazione,
            cellulare = cellulare,
            note = note,
            email1 = email1,
            email2 = email2,
            paginaweb = paginaWeb
        ))
        print("tutto bene")
        id = ris.inserted_primary_key[0]
    except Exception as error:
        conn.rollback()
        print("rip")
        print(error.__cause__)


    return redirect(url_for('.home', id= id))

@contatto_bp.route('/addContattoFile', methods=['POST'])
@login_required
def addContatti():
    print("sono quya, contatti")
    # Read the File using Flask request
    file = request.files['fileContatti']
 
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(file)
    
    # Define variable to read sheet
    dataframe1 = dataframe.active
    print(dataframe1.max_row)
    try:
        for col in range(1, dataframe1.max_row):
            list = []
            for row in dataframe1.iter_cols(1, dataframe1.max_column):
                list.append(row[col].value)
            try:
                conn.execute(insert(contatto).values(
                    nome = list[0],
                    idutente = current_user.get_id(),
                    secondonome = list[1],
                    cognome = list[2],
                    viaufficio1 = list[3],
                    viaufficio2 = list[4],
                    viaufficio3 = list[5],
                    citta = list[6],
                    provincia = list[7],
                    cap = list[8],
                    #numUfficio = list[9],
                    #numUfficio2 = list[10],
                    telefonoprincipale = filterNumber(list[11]),
                    faxabitazione = filterNumber(list[12]),
                    abitazione = list[13],
                    abitazione2 = list[14],
                    cellulare = filterNumber(list[15]),
                    note = list[16],
                    paginaweb = list[17],
                    email1 = list[18],
                    email2 = list[19]
                ))
            
            except Exception as error:
                print("ERRROREEEEEEEEEEEEEEEEEEEEEEE")
                print(error.__cause__)
        print("okokoko")
        
    except Exception as error:
        print("rip")
        print(error.__cause__)
        conn.rollback()
    
    return redirect(url_for('.home', id = 0))

@contatto_bp.route('/modifyContatto/<int:id>', methods=['POST'])
@login_required
def modifyCliente(id):
    print("modifica conattto")
    print(request.form)
    nome = request.form['nome']
    secondoNome = request.form['secondoNome']
    cognome = request.form['cognome']
    viaUfficio1 = request.form['viaUfficio1'] 
    viaUfficio2 = request.form['viaUfficio2']   
    viaUfficio3 = request.form['viaUfficio3']
    provincia = request.form['provincia']
    cap = request.form['cap']
    numUfficio = request.form['numUfficio']
    numUfficio2 = request.form['numUfficio2']
    telefonoPrincipale = request.form['telefonoPrincipale']
    faxAbitazione = request.form['faxAbitazione']
    abitazione = request.form['abitazione']
    cellulare = request.form['cellulare']
    note = request.form['note']
    email1 = request.form['email1']
    email2 = request.form['email2']
    paginaWeb = request.form['paginaWeb']

    try:
        conn.execute(update(contatto).where(contatto.c.idcontatto == id).values(
            nome = nome,
            idutente = current_user.get_id(),
            secondonome = secondoNome,
            cognome = cognome,
            viaufficio1 = viaUfficio1,
            viaufficio2 = viaUfficio2,
            viaufficio3 = viaUfficio3,
            provincia = provincia,
            cap = cap,
            numufficio = numUfficio,
            numufficio2 = numUfficio2,
            telefonoprincipale = telefonoPrincipale,
            faxabitazione = faxAbitazione,
            abitazione = abitazione,
            cellulare = cellulare,
            note = note,
            email1 = email1,
            email2 = email2,
            paginaweb = paginaWeb
        ))
        print("tutto bene")
    except Exception as error:
        print("rip")
        print(error.__cause__)
        conn.rollback()

    return redirect(url_for('.home',id = id))

@contatto_bp.route('/remove/<int:id>', methods=['POST'])
@login_required
def removeContatto(id):
    print("ok")
    try:
        print(id)
        conn.execute(delete(contatto).where(contatto.c.idcontatto == id))
    except Exception as error:
        print("rip")
        print(error.__cause__)
        conn.rollback()
    return url_for('.home', id=0)