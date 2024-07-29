from flask import Blueprint,render_template, url_for, redirect, request, Response
from home import *
from flask_login import *
from model import *
from login.login import bcrypt
import openpyxl
from datetime import date
from datetime import datetime
import io
import xlwt
import psycopg2
import psycopg2.extras
import warnings

warnings.simplefilter(action='ignore', category=UserWarning)

titolo = "Home"
@home_bp.route('/', methods=['GET', 'POST'])
@login_required
def home():
    contatti = conn.execute(select(contatto).where(contatto.c.idutente == current_user.get_id())).fetchall()
    
    """conn.execute(insert(utente).values(
        nome = 'Marco',
        cognome= 'Piazzon',
        email= 'mp@gmail.com',
        psw=bcrypt.generate_password_hash('mp@gmail.com'+'ciao').decode('utf-8'),
        ))
    conn.commit()"""
    getPortafogliUtente = conn.execute(select(portafoglio).where(portafoglio.c.idutente == current_user.get_id()).order_by(portafoglio.c.idportafoglio.desc())).fetchall()
    getPortafogliUtente = list(getPortafogliUtente)
    print(getPortafogliUtente)
    print(type(getPortafogliUtente))
    print("ho fatto questo")
    return render_template ("/home/home.html", getP=getPortafogliUtente, len=len(getPortafogliUtente), titolo=titolo)

@home_bp.route("/delete/<int:id>", methods=['POST','GET'])
@login_required
def removePortafoglio(id):
    print("rimuovo")
    print("sto cancellando")
    print(id)
    try:
        #Cancello tabella appuntamento 
        # DELETE appuntamento FROM appuntamento join trattativaappuntamento on appuntamento.idAppuntamento = trattativaappuntamento.idAppuntamento 
        # join trattativa on trattativa.idTrattativa = trattativaappuntamento.idTrattativa join cliente on trattativa.idCliente = cliente.idCliente where cliente.idPortafoglio = 6;
        conn.execute(delete(appuntamento).\
                                    where(appuntamento.c.idappuntamento == trattativaappuntamento.c.idappuntamento).\
                                    where(trattativaappuntamento.c.idtrattativa == trattativa.c.idtrattativa).\
                                    where(trattativa.c.idcliente == cliente.c.idcliente).\
                                    where(cliente.c.idportafoglio == id)                                                    
        )

        #Cancello tabella trattativaappuntamento


        conn.execute(delete(trattativaappuntamento).\
                where(trattativa.c.idcliente == cliente.c.idcliente).\
                where(trattativaappuntamento.c.idtrattativa == trattativa.c.idtrattativa).\
                where(trattativa.c.idcliente == cliente.c.idcliente).\
                where(cliente.c.idportafoglio == id)
        )
        

        #Cancello trattattive
        # delete trattativa from trattativa join cliente on trattativa.idCliente = cliente.idCliente where cliente.idPortafoglio = 6;
        conn.execute(delete(trattativa).\
                    where(trattativa.c.idcliente == cliente.c.idcliente).\
                    where(cliente.c.idportafoglio == id)
        )
        
        #Cancello clienti
        conn.execute(delete(cliente).where(cliente.c.idportafoglio == id))

        #Cancello portafoglio

        conn.execute(delete(portafoglio).where(portafoglio.c.idportafoglio == id))
        
        print("ho cancellato quello che dovevo fare")
    except Exception as error:
        print("rip")
        print(error.__cause__)
        print(error)
        conn.rollback()


    return redirect(url_for('home_bp.home'))


@home_bp.route('/portafoglio/<int:id>')
@login_required
def goToPortafoglio(id, message):
    print("cviao")
    print(id)
    clienti = conn.execute(select(cliente).where(cliente.c.idportafoglio == id)).fetchall()
    return render_template("/portafoglio/portafoglio.html", clienti=clienti, message = message, titolo ="Portafoglio")

def filterNumber(val):
    if(val is None):
        return None
    
    if(val is String):
        sub = sub.replace(" ", "")
        sub = list(val)
        if(len(sub) == 13 and list(val)[0] == '+'): # esempio '+393453659562' --> 3453659562
            return val[2:]
    
    return str(val)


def getFase(andtra, value):
    print(type(andtra[-1][0]))
    if(value is None):
        return andtra[-1][0]
    for v in andtra:
        if(value.upper() == v.nome.upper()):
            return v.idandamento
    return andtra[-1][0]

def getCategoria(getCateOff, value):
    print(getCateOff[-1][0])
    if(value is None):
        return getCateOff[-1][0]
    for v in getCateOff:
        if(value.upper() == v.nome.upper()):
            print("trovato" + str(v[0]))
            return v.idcategoria
    return getCateOff[-1][0]
"""
def convertDate(val):
    if(val is None):
        return None
    
    print
    split_val = date.strftime(val, "%").split("-")
    if(len(split_val) != 2):
        return None
     
    months = ["gennaio","febbraio","marzo","aprile","maggio","giugno","luglio","agosto","settembre","ottobre","novembre","dicembre"]
    month = None
    
    for i in range(len(months)):
        if(months[i] == split_val[0]):
            month = i+1
    
    day = str(split_val[1])
    if(month is None or day is None):
        return None
    
    return date(date.year, month, day) 
"""

def setZero(val):
    if val is None:
        return 0
    else: 
        return val

@home_bp.route('/addPortafoglio',methods=['POST'])
@login_required
def addPortafoglio():
    warnings.simplefilter(action='ignore', category=UserWarning)
    print("sono quya")
    # Read the File using Flask request
    file = request.files['fileDoc']
 
    # Parse the data as a Pandas DataFrame type
    #data = pandas.read_excel(file)
 
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(file)
    
    # Define variable to read sheet
    dataframe1 = dataframe.active

    file2 = request.files['filePipeline']
    # Parse the data as a Pandas DataFrame type
    #data = pandas.read_excel(file)

    newid = 0
    try:
        res = conn.execute(insert(portafoglio).values(
            idutente = current_user.get_id(),
            datainserimento = date.today()
        ))

        newid = res.inserted_primary_key[0]
        
        current_user.idport = newid
        #print("test della vita")
        print("test id")
        print(newid)
        print(res.lastrowid)
        for col in range(1, dataframe1.max_row):
            listapp = []
            for row in dataframe1.iter_cols(1, dataframe1.max_column):
                listapp.append(row[col].value)
           # print("sono qua 3")
           # print(listapp)
            conn.execute(insert(cliente).values(
                idutente = current_user.get_id(),
                idportafoglio = newid,
                tipocliente = conn.execute(select(tipocliente.c.idtipocliente).where(tipocliente.c.nome == listapp[0])).fetchone()[0], #da testare
                cf = listapp[1],
                ragionesociale = listapp[2],
                
                presidio = listapp[3],
                indirizzoprincipale = (listapp[4]),
                comuneprincipale = listapp[5],
                capprincipale = listapp[6],
                provinciadescprincipale = listapp[7], 
                provinciasiglaprincipale = listapp[8],
                seditot = setZero(listapp[9]),
                nlineetot = setZero(listapp[10]),
                fisso = setZero(listapp[11]),
                mobile = setZero(listapp[12]),
                totale = setZero(listapp[13]),
                fatturatocerved = setZero(listapp[14]),
                clienteoffmobscadenza = (listapp[15]),
                fatturatotim = setZero(listapp[16]),
                dipendenti = setZero(listapp[17])
            ))
            print("insert")
        print("okokoko")
        print("sono qua 1")
        if not (file2 is None):
             
            # Define variable to load the dataframe
            dataframe2 = openpyxl.load_workbook(file2, data_only=True)
            
            # Define variable to read sheet
            dataframe3 = dataframe2.active
            #print(dataframe3.max_row)
            andtra = conn.execute(select(andamentotrattativa)).fetchall()

            #print(andtra)
            #print(andtra[0][0])
            #print(type(andtra[0][0]))
            
            getCateOff = conn.execute(select(categoria)).fetchall()
            print("dopo")
            for col in range(1, dataframe3.max_row):
                listapp = []
                checkRow = True
                for row in dataframe3.iter_cols(1, dataframe3.max_column):
                    if(row[col].value is not None):
                        checkRow = False
                    #print((row[col].value), end= ", ")
                    listapp.append(row[col].value)
                
                if(checkRow == False):
                    #print("sto testando id")
                    #print(listapp)
                    #print(listapp[3])
                    idlist = conn.execute(select(cliente.c.idcliente,cliente.c.idportafoglio).where(listapp[3] == cliente.c.ragionesociale)).fetchall()
                    print(len(idlist))
                    #print(idlist)
                    idlist = list(idlist)
                    print("lista")
                    print(idlist)
                    numero = idlist[-1]
                    prob = None
                    if(listapp[19]):
                        prob = listapp[19]*100
                    try:
                        conn.execute(insert(trattativa).values(
                            idutente = current_user.get_id(),
                            idcliente = numero[0], #fare query che trova il nome e assegna l'id nella tabella cliente
                            codicectrdigitali = listapp[0],
                            codicesaleshub = listapp[1],
                            areamanager = conn.execute(select(utente.c.areamanager).where(utente.c.idutente == current_user.get_id())).fetchone()[0],
                            zona = listapp[4],
                            tipo = (listapp[5]),
                            nomeopportunita = listapp[6],
                            datacreazioneopportunita = (listapp[7]), # da testare
                            fix = setZero(listapp[8]),
                            mobile = setZero(listapp[9]),
                            categoriaoffertait = getCategoria(getCateOff, listapp[10]),
                            it = setZero(listapp[11]),
                            lineefoniafix = setZero(listapp[12]),
                            aom = setZero(listapp[13]),
                            mnp = setZero(listapp[14]),
                            al = setZero(listapp[15]),
                            datachiusura = (listapp[16]),
                            fase = getFase(andtra, listapp[17]) ,
                            notespecialista = (listapp[18]),
                            probabilita = prob,
                            inpaf = None,
                            fornitore = listapp[22],
                        ))
                    except Exception as error:
                        print("rip row interna")
                        print(error)
                        print(error.__cause__)
        print("okokoko")
        global message
        message = "Portafoglio creato"
    except Exception as error:
        print("rip")
        print(error)
        print(error.__cause__)
    
    return redirect(url_for("portafoglio_bp.home", idPort = newid, id = 0))


@home_bp.route('/getExcel/<int:id>')
def getExcel(id):
    print("sto creando il file")
    print(id)
    res = conn.execute(select(cliente).where(cliente.c.idportafoglio == id)).fetchall()
    output = io.BytesIO()

    workbook = xlwt.Workbook()

    sh = workbook.add_sheet('Report Portafoglio')

    sh.write(0,0,'TIPO CLIENTE')
    sh.write(0,1,'CF')
    sh.write(0,2,'RAG_SOC_CLI')
    sh.write(0,3,'PRESIDIO')
    sh.write(0,4,'INDIRIZZO PRINCIPALE')
    sh.write(0,5,'COMUNE_PRINCIPALE')
    sh.write(0,6,'CAP_PRINCIPALE')
    sh.write(0,7,'PROVINCIA_DESCR_PRINCIPALE')
    sh.write(0,8,'PROVINCIA_SIGLA_PRINCIPALE')
    sh.write(0,9,'SEDI_TOT')
    sh.write(0,10,'N_LINEE_TOT')
    sh.write(0,11,'_FISSO')
    sh.write(0,12,'_MOBILE')
    sh.write(0,13,'_TOTALE')
    sh.write(0,14,'FATTURATO_CERVED')
    sh.write(0,15,'CLIENTE_OFF_MOB_SCADENZA')
    sh.write(0,16,'FATTURATO_IT_TIM')
    sh.write(0,17,'DIPENDENTI')

    idx = 0
    for row in res:
        sh.write(idx+1, 0, row[3])
        sh.write(idx+1, 1, row[4])
        sh.write(idx+1, 2, row[5])
        sh.write(idx+1, 3, row[6])
        sh.write(idx+1, 4, row[7])
        sh.write(idx+1, 5, row[8])
        sh.write(idx+1, 6, row[9])
        sh.write(idx+1, 7, row[10])
        sh.write(idx+1, 8, row[11])
        sh.write(idx+1, 9, row[12])
        sh.write(idx+1, 10, row[13])
        sh.write(idx+1, 11, row[14])
        sh.write(idx+1, 12, row[15])
        sh.write(idx+1, 13, row[16])
        sh.write(idx+1, 14, row[17])
        sh.write(idx+1, 15, row[18])
        sh.write(idx+1, 16, row[19])
        sh.write(idx+1, 17, row[20])
        idx += 1
    
    workbook.save(output)
    output.seek(0)
    
    return Response(output, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=report_portafoglio.xls"})
        
""""
  idUtente = current_user.get_id(),
                idPortafoglio = lastPortafoglio,
                tipoCliente = 1,
                cf = correctValue(list[1],
                ragioneSociale = correctValue(list[2]),
                presidio = 1,
                indirizzoPrincipale = correctValue(list[4]),
                comunePrincipale = 1,
                capPrincipale = correctValue(list[6]),
                provinciaDescPrincipale = correctValue(list[7]),
                provinciaSiglaPrincipale = correctValue(list[8]),
                sediTot = correctValue(list[9]),
                nLineeTot = correctValue(list[10]),
                fisso = correctValue(list[11]),
                mobile = correctValue(list[12]),
                totale = correctValue(list[12]),
                fatturatoCerved = correctValue(list[13]),
                clienteOffMobScadenza = correctValue(list[14]),
                fatturatoTim = correctValue(list[15]),
                dipendenti = correctValue(list[16])"""
